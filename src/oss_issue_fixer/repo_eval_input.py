from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

REPO_URL_HEADERS = (
    "仓库链接",
    "倉庫鏈接",
    "꾑욋졍쌈",
    "repo_url",
    "repository_url",
    "url",
)


def load_repos_from_xlsx(
    path: str,
    sheet_name: str = "",
) -> list[str]:
    workbook = load_workbook(Path(path), data_only=True, read_only=True)
    worksheet = workbook[sheet_name] if sheet_name else workbook[workbook.sheetnames[0]]
    rows = worksheet.iter_rows(values_only=True)
    header = next(rows, None)
    if not header:
        return []

    header_map: dict[str, int] = {}
    for index, raw in enumerate(header):
        key = str(raw or "").strip()
        if key:
            header_map[key] = index

    repo_col = None
    for candidate in REPO_URL_HEADERS:
        if candidate in header_map:
            repo_col = header_map[candidate]
            break
    if repo_col is None:
        raise ValueError(
            "repo url column not found in xlsx header, expected one of: "
            + ", ".join(REPO_URL_HEADERS)
        )

    repos: list[str] = []
    seen: set[str] = set()
    for row in rows:
        if row is None:
            continue
        value = row[repo_col] if repo_col < len(row) else None
        repo = str(value or "").strip()
        if not repo or repo in seen:
            continue
        seen.add(repo)
        repos.append(repo)
    return repos
